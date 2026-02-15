from io import BytesIO
from datetime import date, datetime, time
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


def export_au12_excel(rs) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "АУ-12"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def v(val):
        if val in [None, '']:
          return "-"
        if isinstance(val, (str, int, float, bool, Decimal, date, datetime, time)):
          return val
        return str(val)

    def cell(r, c, value="", bold=False, wrap=True, align="left"):
        ce = ws.cell(row=r, column=c, value=value)
        ce.font = Font(bold=bold)
        ce.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
        ce.border = border
        return ce

    # ширины
    for i, w in enumerate([12, 24, 24, 12, 12, 24, 24, 14, 14, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # заголовок
    r = 1
    ws.merge_cells("A1:K1")
    cell(1, 1, "МАРШРУТНЫЙ ЛИСТ формы АУ-12", bold=True, align="center")
    r += 1
    ws.merge_cells("A2:K2")
    cell(2, 1, f"№ {rs.number} от {rs.date}", bold=True, align="center")

    r += 1
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, f"Дорога: {v(rs.railway_name)}")
    r += 1
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, f"Предприятие: {rs.enterprise.name if rs.enterprise else '—'}")
    r += 1
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, f"ССПС: {rs.ssps_unit}")
    r += 1
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, f"Бригада: {rs.brigade.name if rs.brigade else '—'}")

    r += 2
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, "Раздел I. Бригада", bold=True)
    r += 1
    headers1 = ["Должность", "ФИО", "Явка", "Оконч.", "Перер.", "Отдых"]
    for col, h in enumerate(headers1, 1):
        cell(r, col, h, bold=True, align="center")
    r += 1


    # Раздел II
    
    for row in rs.crew_rows.all():
        cell(r, 1, v(row.position_display))
        cell(r, 2, v(row.full_name))
        cell(r, 3, v(row.time_check_in), align="center")
        cell(r, 4, v(row.time_check_out), align="center")
        cell(r, 5, v(row.overtime_minutes), align="center")
        cell(r, 6, v(row.rest_minutes), align="center")
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, "Раздел II. Пробег и ТСМ", bold=True)
    r += 1
    s2 = getattr(rs, "section_ii", None)
    owner_name = "—"
    if s2 and s2.transport and s2.transport.ssps_unit and s2.transport.ssps_unit.owner:
        owner_name = s2.transport.ssps_unit.owner.name
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, f"Предприятие-владелец: {owner_name}")
    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    cell(r, 1, f"Тип машины: {s2.machine_type.name if s2 and s2.machine_type else '—'}")
    ws.merge_cells(f"F{r}:K{r}")
    cell(r, 6, f"Номер: {s2.transport.number if s2 and s2.transport else '—'}")
    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    cell(r, 1, f"Спидометр: {v(s2.speedometer_out_km) if s2 else '—'} / {v(s2.speedometer_in_km) if s2 else '—'}")
    ws.merge_cells(f"F{r}:K{r}")
    cell(r, 6, f"Моточасы: {v(s2.moto_hours_out) if s2 else '—'} / {v(s2.moto_hours_in) if s2 else '—'}")
    r += 1
    ws.merge_cells(f"A{r}:K{r}")
    cell(
        r,
        1,
        f"Топливо: {v(s2.fuel_type) if s2 else '—'} | Выдано: {v(s2.fuel_issued_l) if s2 else '—'} | "
        f"Ост. выезд: {v(s2.fuel_left_out_l) if s2 else '—'} | Ост. возврат: {v(s2.fuel_left_in_l) if s2 else '—'}",
    )
    
    r += 2

    # Раздел III
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, "Раздел III. Работы", bold=True)
    r += 1
    headers3 = ["№", "Станция от", "Станция до", "Отпр.", "Приб.", "Работы", "Место"]    
    for col, h in enumerate(headers3, 1):
        cell(r, col, h, bold=True, align="center")
    r += 1

    for w in rs.work_rows.all():
        cell(r, 1, v(w.request_no))
        cell(r, 2, v(w.station_from))
        cell(r, 3, v(w.station_to))
        cell(r, 4, v(w.time_departure), align="center")
        cell(r, 5, v(w.time_arrival), align="center")
        cell(r, 6, v(w.work_name))
        cell(r, 7, v(w.work_place))
        r += 1


    s4 = getattr(rs, "section_iv", None)
    s5 = getattr(rs, "section_v", None)
    s6 = getattr(rs, "section_vi", None)
    
    r += 2
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, "Раздел IV — результаты работы и расход ТСМ", bold=True)
    r += 1
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, f"Состояние: {v(s4.technical_state) if s4 else '—'}")
    r += 1
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, f"Неисправности: {v(s4.defects_found) if s4 else '—'}")
    r += 1
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, f"Меры: {v(s4.measures_taken) if s4 else '—'}")

    r += 1
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, "Раздел V — тех. состояние и допуски", bold=True)
    r += 1
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, f"Допуски/примечания: {v(s5.fuel_notes) if s5 else '—'}")

    r += 1
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, f"Связь/устройства: {v(s5.repairs_notes) if s5 else '—'}")
    r += 1
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, f"Ответственный: {v(s5.master_name) if s5 else '—'}")

    r += 1
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, "Раздел VI — замечания", bold=True)
    r += 1
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, f"Замечания: {v(s6.final_notes) if s6 else '—'}")
    r += 1
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, f"Проверил: {v(s6.inspector_name) if s6 else '—'}")

    r += 2
    med = getattr(rs, "medical", None)
    ws.merge_cells(f"A{r}:K{r}")
    cell(r, 1, "Медосмотр", bold=True)
    r += 1
    if med:
        ws.merge_cells(f"A{r}:K{r}")
        cell(r, 1, f"До: {v(med.pre_time)} · допуск: {'ДА' if med.pre_allowed else 'НЕТ'} · медик: {v(med.pre_medic_name)}")
        r += 1
        ws.merge_cells(f"A{r}:K{r}")
        cell(r, 1, f"После: {v(med.post_time)} · допуск: {'ДА' if med.post_allowed else 'НЕТ'} · медик: {v(med.post_medic_name)}")
    else:
        ws.merge_cells(f"A{r}:K{r}")
        cell(r, 1, "Нет данных медосмотра.")
        
    out = BytesIO()
    wb.save(out)
    return out.getvalue()
